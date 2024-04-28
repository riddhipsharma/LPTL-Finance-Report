

# Create your views here.
from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
# 1 -> lastmonth sheet
# 2 -> currentmonth sheet
def post_list(request):
    if "GET" == request.method:
        return render(request, 'dummyCostCode/post_list.html', {})
    else:
        # print ("Hello World") 
        excel_file1 = request.FILES["excel_file1"]  
        excel_file2 = request.FILES["excel_file2"]
        wb1 = load_workbook(excel_file1)
        wb2 = load_workbook(excel_file2)
        # wb = Workbook()
        # ws = wb.create_sheet("Sheet1",0)
        ws1 = wb1['Sheet1']
        ws2 = wb2['Sheet1']

        ws2.insert_cols(idx=3)
        char = get_column_letter(3)
        ws2[char + str(1)].value = "ACTUAL_COST_CENTER"
        i=1
        # col is A B C D ...
        for row1 in range(2, ws1.max_row+1):
            for col1 in range(4,5):
                # print(i)
                char1 = get_column_letter(col1)
                # print(ws1[char1 + str(row1)])
                s1=ws1[char1 + str(row1)].value
                for row4 in range(2,ws2.max_row+1):
                    for col4 in range(4,5):
                        char4 = get_column_letter(col4)
                        # print(ws2[char4 + str(row4)])
                        s2=ws2[char4 + str(row4)].value
                        if(s1==s2):
                            # print("same sample number")

                            for col2 in range(8,9):
                                char2 = get_column_letter(col2)
                                # print(ws1[char2 + str(row1)])
                                # print(ws2[char2 + str(row4)])
                                t1=ws1[char2 + str(row1)].value
                                t2=ws2[char2 + str(row4)].value
                                if t1==t2:
                                    # print("same test code")
                                    for col3 in range(3,4):
                                        char3 = get_column_letter(col3)
                                        # print(ws2[char3 + str(row1)])
                                        # print("charge code updated")
                                        # print()
                                        # ws[char3 + str(row4)].value=ws1[char3 + str(row1)].value
                                        ws2[char3 + str(row4)].value=ws1[char3 + str(row1)].value
                                        # wb2.save("excel_file2")
                                
                        else:
                            # print("Different sample number")
                            continue             
                i=i+1
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=DummyCostCenter.xlsx'
        wb2.save(response)
        return response
    
def sampleResultsAllDates(request):
    if "GET" == request.method:
        return render(request, 'dummyCostCode/sampleResultsAllDates.html', {})
    else:
        excel_file1 = request.FILES["excel_file1"]  
        excel_file2 = request.FILES["excel_file2"]
        # wb1 = load_workbook('./sampleResultsAllDates/QTS_LPTL_SAMPLES_RESULTS_ALL_DATES.xlsx')
        # wb2 = load_workbook('./sampleResultsAllDates/QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx')
        wb1 = load_workbook(excel_file1)
        wb2 = load_workbook(excel_file2)
        # wb = Workbook()
        # ws = wb.create_sheet("Sheet1",0)
        ws1 = wb1['Sheet1']
        ws2 = wb2['Sheet1']

        if ws1['AE' + str(1)].value != "BILLED_Y/N":
            ws1.insert_cols(idx=31)
            char = get_column_letter(31)
            ws1[char + str(1)].value = "BILLED_Y/N"
            # wb1.save('./sampleResultsAllDates/QTS_LPTL_SAMPLES_RESULTS_ALL_DATES.xlsx')

        # col is A B C D...
        # row is 1 2 3 4...
        for row1 in range(2,ws1.max_row+1):
            # for col1 in range(3,4):
                # print(i)
                # char1 = get_column_letter(col1)
                s1 = ws1['C' + str(row1)].value
                # print(s1)
                for row4 in range(2,ws2.max_row+1):
                    # for col4 in range(3,4):  
                        # char4 = get_column_letter(col4)
                        s2 = ws2['C' + str(row4)].value
                        # print(s2)
                        if(s1==s2 and s1!=None):
                            # print("same sample number"+s1+" "+s2)
                            # for col2 in range(30,31):
                            #     char2 = get_column_letter(col2)
                                t1 = ws1['AC' + str(row1)].value
                                t2 = ws2['G' + str(row4)].value

                                if t1 == t2 and t1!=None:
                                    print(t1)
                                    print(t2)
                                    print("same test code")
                                    ws1['AE' + str(row1)].value = "Y"
                                    print(ws1['AE' + str(row1)].value)
                                    print()
                

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=sampleResultsAllDates.xlsx'
        wb1.save(response)
        return response
        # wb1.save('./sampleResultsAllDates/QTS_LPTL_SAMPLES_RESULTS_ALL_DATES.xlsx')

def calculateTotalBillingAmountDetail(request):
    if "GET" == request.method:
        return render(request, 'dummyCostCode/calculateTotalBillingAmountDetail.html', {})
    else:
        # QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW
        excel_file1 = request.FILES["excel_file1"]  
        wb1 = load_workbook(excel_file1)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx')
        ws1 = wb1['Sheet1']
        first_row = 2
        last_row = ws1.max_row
        sum_row = last_row + 3
        start_col = 11
        end_col = 11
        for row in ws1.iter_rows(min_row=sum_row, max_row=sum_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell_sum_start = 'K' + str(first_row)
                cell_sum_end = 'K' + str(last_row)
                cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)
        detailall=cell.value
        print("done")

        # wb1.save("./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx")

        # QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW
        excel_file2 = request.FILES["excel_file2"]  
        wb1 = load_workbook(excel_file2)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW.xlsx')
        ws1 = wb1['Sheet1']
        first_row = 2
        last_row = ws1.max_row
        sum_row = last_row + 3
        start_col = 12
        end_col = 12
        for row in ws1.iter_rows(min_row=sum_row, max_row=sum_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell_sum_start = 'L' + str(first_row)
                cell_sum_end = 'L' + str(last_row)
                cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)
        detail=cell.value
        print("done")

        # wb1.save("./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW.xlsx")

        # QTS_LPTL_COSTCENTER_BILLING_DET_CHANGES_VW
        excel_file3 = request.FILES["excel_file3"]  
        wb1 = load_workbook(excel_file3)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')
        ws1 = wb1['Sheet1']
        first_row = 2
        last_row = ws1.max_row
        sum_row = last_row + 3
        start_col = 11
        end_col = 11
        for row in ws1.iter_rows(min_row=sum_row, max_row=sum_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell_sum_start = 'K' + str(first_row)
                cell_sum_end = 'K' + str(last_row)
                cell.value = '=SUM({0}:{1})'.format(cell_sum_start, cell_sum_end)
        detailchanges=cell.value
        print("done")

        # wb1.save('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')

        # filepath = "./finalReportGeneration/test.xlsx"
        wb = Workbook('summationdetails.xlsx')
        ws = wb.create_sheet()
        # wb = Workbook()
        # wb.save(filepath)
        # ws=wb['Sheet']
        # ws.append(row)
        # ws.append(row)
        ws['A2']="Detail All"
        ws['A2']="Detail"
        ws['A3']="Detail Changes"
        ws['B1'].value=detailall
        ws['B2'].value=detail
        ws['B3'].value=detailchanges
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=summationdetails.xlsx'
        wb.save(response)
        return response

# def calculateTotalBillingAmountSummary(request):
#     if "GET" == request.method:
#          return render(request, 'dummyCostCode/compSumAndDetail.html', {})
#     else:
         
    
def compSumAndDetail(request):
    if "GET" == request.method:
         return render(request, 'dummyCostCode/compSumAndDetail.html', {})
    else:
        #QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx
        excel_file1 = request.FILES["excel_file1"]  
        wb1 = load_workbook(excel_file1, data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(11)
        last_row = ws1.max_row
        detAll=ws1[char+str(last_row)].value

        # QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW
        excel_file2 = request.FILES["excel_file2"]
        wb1 = load_workbook(excel_file2, data_only=True)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW.xlsx', data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(12)
        last_row = ws1.max_row
        det=ws1[char+str(last_row)].value

        # QTS_LPTL_COSTCENTER_BILLING_DET_CHANGES_VW
        excel_file3 = request.FILES["excel_file3"]
        wb1 = load_workbook(excel_file3, data_only=True)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx', data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(11)
        last_row = ws1.max_row
        # print(last_row)
        # print(ws1[char+str(last_row)].value)
        detChanges=ws1[char+str(last_row)].value

        # QTS_LPTL_COSTCENTER_BILLING_SUM_ALL_VW
        excel_file4 = request.FILES["excel_file4"]
        wb1 = load_workbook(excel_file4, data_only=True)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUM_ALL_VW.xlsx', data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(5)
        last_row = ws1.max_row
        print(ws1[char+str(last_row)].value)
        sumAll=ws1[char+str(last_row)].value

        # QTS_LPTL_COSTCENTER_BILLING_SUMMARY_VW
        excel_file5 = request.FILES["excel_file5"]
        wb1 = load_workbook(excel_file5, data_only=True)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUMMARY_VW.xlsx', data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(5)
        last_row = ws1.max_row
        print(ws1[char+str(last_row)].value)
        sum=ws1[char+str(last_row)].value

        # QTS_LPTL_COSTCENTER_BILL_SUM_CHANGES_VW
        excel_file6 = request.FILES["excel_file6"]
        wb1 = load_workbook(excel_file6, data_only=True)
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_SUM_CHANGES_VW.xlsx', data_only=True)
        ws1 = wb1['Sheet1']
        char = get_column_letter(5)
        last_row = ws1.max_row
        # print(last_row)
        print(ws1[char+str(last_row)].value)
        sumChanges=ws1[char+str(last_row)].value
        context = {
            'message1': 'CORRECT',
            'message2': 'INCORRECT',
        }
        if detAll==sumAll and det==sum and detChanges==sumChanges:
            return render(request, 'dummyCostCode/correct.html', context)
        else:
            return render(request, 'dummyCostCode/incorrect.html', context)
        

def compPrevDetChangesAndCurDetAll(request):
    if "GET" == request.method:
         return render(request, 'dummyCostCode/compPrevDetChangesAndCurDetAll.html', {})
    else:
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')
        # wb2 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx')
        # wb3 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW.xlsx')
        # wb4 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUM_ALL_VW.xlsx')
        # wb5 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUMMARY_VW.xlsx')
        # wb6 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_SUM_CHANGES_VW.xlsx')
        # wb7 = load_workbook('./finalReportGeneration/QTS_LPTL_SAMPLES_RESULTS_ALL_DATES.xlsx')
        excel_file1 = request.FILES["excel_file1"]  
        wb1 = load_workbook(excel_file1, data_only=True)

        excel_file2 = request.FILES["excel_file2"]  
        wb2 = load_workbook(excel_file2, data_only=True)

        # excel_file3 = request.FILES["excel_file3"]  
        # wb3 = load_workbook(excel_file3, data_only=True)

        # excel_file4 = request.FILES["excel_file4"]  
        # wb4 = load_workbook(excel_file4, data_only=True)

        # excel_file5 = request.FILES["excel_file5"]  
        # wb5 = load_workbook(excel_file5, data_only=True)

        # excel_file6 = request.FILES["excel_file6"]  
        # wb6= load_workbook(excel_file6, data_only=True)

        # excel_file7 = request.FILES["excel_file7"]  
        # wb7 = load_workbook(excel_file7, data_only=True)

        ws1 = wb1['Sheet1']     #QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW
        ws2 = wb2['Sheet1']     #QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW
        # ws3 = wb3['Sheet1']
        # ws4 = wb4['Sheet1']
        # ws5 = wb5['Sheet1']
        # ws6 = wb6['Sheet1']
        # ws7 = wb7['Sheet1']

        # comparing current month's detail changes with previous month's detail all to check which samples are already billed

        if ws1['L' + str(1)].value != "BILLED_Y/N":
            ws1.insert_cols(idx=12)
            char = get_column_letter(12)
            ws1[char + str(1)].value = "BILLED_Y/N"
            # wb1.save('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')

        last_row_dc = ws1.max_row 
        print(last_row_dc)
        last_row_da = ws2.max_row 

        for row1 in range(2,last_row_dc):
            # for col1 in range(3,4):
                # print(i)
                # char1 = get_column_letter(col1)
                s1 = ws1['C' + str(row1)].value
                # print(s1)
                for row4 in range(2,last_row_da):
                    # for col4 in range(3,4):  
                        # char4 = get_column_letter(col4)
                        s2 = ws2['C' + str(row4)].value
                        # print(s2)
                        if(s1==s2 and s1!=None):
                            # print("same sample number"+s1+" "+s2)
                            # for col2 in range(30,31):
                            #     char2 = get_column_letter(col2)
                                t1 = ws1['G' + str(row1)].value
                                t2 = ws2['G' + str(row4)].value

                                if t1 == t2 and t1!=None:
                                    print(t1)
                                    print(t2)
                                    print("same test code")
                                    ws1['L' + str(row1)].value = "Y"
                                

        # wb1.save('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')

        allBilled=False
        someBilled=False
        count=0
        for row in range(2,last_row_dc-2):
            if ws1['L' + str(row)].value == 'Y':
                    count=count+1

        if count==last_row_dc:
            allBilled=True
        elif count==0:
            allBilled=False
        elif count!=0 and count<last_row_dc:
            someBilled=True

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx'
        wb1.save(response)
        return response
    
def finalReport(request):
    if "GET" == request.method:
         return render(request, 'dummyCostCode/finalReport.html', {})
    else:
        # wb1 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW.xlsx')
        # wb2 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW.xlsx')
        # wb3 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW.xlsx')
        # wb4 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUM_ALL_VW.xlsx')
        # wb5 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILLING_SUMMARY_VW.xlsx')
        # wb6 = load_workbook('./finalReportGeneration/QTS_LPTL_COSTCENTER_BILL_SUM_CHANGES_VW.xlsx')
        # wb7 = load_workbook('./finalReportGeneration/QTS_LPTL_SAMPLES_RESULTS_ALL_DATES.xlsx')
        excel_file1 = request.FILES["excel_file1"]  
        wb1 = load_workbook(excel_file1, data_only=True)

        excel_file2 = request.FILES["excel_file2"]  
        wb2 = load_workbook(excel_file2, data_only=True)

        excel_file3 = request.FILES["excel_file3"]  
        wb3 = load_workbook(excel_file3, data_only=True)

        excel_file4 = request.FILES["excel_file4"]  
        wb4 = load_workbook(excel_file4, data_only=True)

        excel_file5 = request.FILES["excel_file5"]  
        wb5 = load_workbook(excel_file5, data_only=True)

        excel_file6 = request.FILES["excel_file6"]  
        wb6= load_workbook(excel_file6, data_only=True)

        excel_file7 = request.FILES["excel_file7"]  
        wb7 = load_workbook(excel_file7, data_only=True)

        excel_file8 = request.FILES["excel_file8"]  
        wb = load_workbook(excel_file8, data_only=True)

        ws1 = wb1['Sheet1']     #QTS_LPTL_COSTCENTER_BILL_DET_CHANGES_VW
        ws2 = wb2['Sheet1']     #QTS_LPTL_COSTCENTER_BILLING_DET_ALL_VW
        ws3 = wb3['Sheet1']     #QTS_LPTL_COSTCENTER_BILLING_DETAIL_VW
        ws4 = wb4['Sheet1']     #QTS_LPTL_COSTCENTER_BILLING_SUM_ALL_VW
        ws5 = wb5['Sheet1']     #QTS_LPTL_COSTCENTER_BILLING_SUMMARY_VW
        ws6 = wb6['Sheet1']     #QTS_LPTL_COSTCENTER_BILL_SUM_CHANGES_VW
        ws7 = wb7['Sheet1']     #QTS_LPTL_SAMPLES_RESULTS_ALL_DATES
        ws = wb['Sheet1']       #LPTL_Transer_Report

        # comparing current month's detail changes with previous month's detail all to check which samples are already billed

        last_row_dc = ws1.max_row 
        print(last_row_dc)
        last_row_da = ws2.max_row 

        allBilled=False
        someBilled=False
        count=0
        for row in range(2,last_row_dc-2):
            if ws1['L' + str(row)].value == 'Y':
                    count=count+1

        if count==last_row_dc:
            allBilled=True
        elif count==0:
            allBilled=False
        elif count!=0 and count<last_row_dc:
            someBilled=True
    
        print(allBilled)
        print(someBilled)

        # wb = Workbook('LPTL_Transer_Report.xlsx')
        # ws = wb.create_sheet()

        # filepath = "./finalReportGeneration/LPTL_Transer_Report.xlsx"
        # wb = Workbook()
        # wb.save(filepath)

        # CASE 01: if all samples in detchanges are billed. Copy sum file as final report

        if(allBilled==True):
            # source_file = excel_file5
            # source_sheet_name = "Sheet1"
            # output_file = wb

            # source_wb = openpyxl.load_workbook(excel_file5)
            source_ws = ws5

            # output_wb = openpyxl.Workbook()
            output_ws = ws
            # output_ws.title = source_sheet_name

            for row in source_ws.iter_rows():
                for cell in row:
                    output_ws[cell.coordinate].value = cell.value
                    output_ws[cell.coordinate].font = cell.font.copy()
                    output_ws[cell.coordinate].border = cell.border.copy()
                    output_ws[cell.coordinate].fill = cell.fill.copy()
                    output_ws[cell.coordinate].number_format = cell.number_format
                    output_ws[cell.coordinate].protection = cell.protection.copy()
                    output_ws[cell.coordinate].alignment = cell.alignment.copy()
                    output_ws[cell.coordinate].comment = cell.comment
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=LPTL_Transer_Report.xlsx'
            wb.save(response)
            return response

        # CASE 02: if not even a single sample in detchanges are billed in detailall. Copy sumall file

        # if allBilled==True:
        #      print("all samples in detchanges are billed in detailall")

        # if allBilled==False:
        #      print("not even a single sample in detchanges are billed in detailall")
                            
        # if someBilled==True:
        #      print("some samples in detchanges are billed in detailall")
            
        if allBilled == False and someBilled == False:
            # source_file = excel_file4
            # source_sheet_name = "Sheet1"
            # output_file = wb

            # source_wb = openpyxl.load_workbook(excel_file4)
            source_ws = ws4

            # output_wb = openpyxl.Workbook()
            output_ws = ws
            # output_ws.title = source_sheet_name

            for row in source_ws.iter_rows():
                for cell in row:
                    output_ws[cell.coordinate].value = cell.value
                    output_ws[cell.coordinate].font = cell.font.copy()
                    output_ws[cell.coordinate].border = cell.border.copy()
                    output_ws[cell.coordinate].fill = cell.fill.copy()
                    output_ws[cell.coordinate].number_format = cell.number_format
                    output_ws[cell.coordinate].protection = cell.protection.copy()
                    output_ws[cell.coordinate].alignment = cell.alignment.copy()
                    output_ws[cell.coordinate].comment = cell.comment
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=LPTL_Transer_Report.xlsx'
            wb.save(response)
            return response

        # Case 03: if some samples in detchanges are billed. Copy sumall file as final report and subtract

        elif allBilled == False and someBilled == True:
            # source_file = wb4
            # source_sheet_name = "Sheet1"
            # output_file = wb

            # source_wb = openpyxl.load_workbook(source_file)
            source_ws = ws4

            # output_wb = openpyxl.Workbook()
            output_ws = ws
            # output_ws.title = source_sheet_name

            for row in source_ws.iter_rows():
                for cell in row:
                    output_ws[cell.coordinate].value = cell.value
                    output_ws[cell.coordinate].font = cell.font.copy()
                    output_ws[cell.coordinate].border = cell.border.copy()
                    output_ws[cell.coordinate].fill = cell.fill.copy()
                    output_ws[cell.coordinate].number_format = cell.number_format
                    output_ws[cell.coordinate].protection = cell.protection.copy()
                    output_ws[cell.coordinate].alignment = cell.alignment.copy()
                    output_ws[cell.coordinate].comment = cell.comment
            # output_wb.save(filepath)

            # to check in detchanges file which sample is billed and has status Y and subtract that much billing amount from final report
            for rowdetchanges in range(2, ws1.max_row):
                if ws1['AE' + str(rowdetchanges)].value == "Y":
                    s1 = ws1['C' + str(rowdetchanges)].value
                    for rowbildall in range(2,ws1.max_row):
                        s2 = ws2['C' + str(rowbildall)].value
                        if s1==s2:
                            t1 = ws1['G' + str(rowdetchanges)].value
                            t2 = ws2['G' + str(rowbildall)].value
                            if t1 == t2:
                                    print(s1+t1)
                                    print(s2+t2)
                                    print("same sample and test code")
                                    billingAmount=ws2['K' + str(rowbildall)].value
                                    costCenter=ws2['B' + str(rowbildall)].value
                                    
                                    for rowfinal in range(2,output_ws.max_row):
                                        #   if prevt1!=t1 and prevs1!=s1:
                                            if output_ws['D'+str(rowfinal)].value == costCenter:
                                                afterSubtract = output_ws['E'+str(rowfinal)].value - billingAmount
                                                output_ws['E'+str(rowfinal)].value = afterSubtract
                                                response = HttpResponse(content_type='application/vnd.ms-excel')
                                                response['Content-Disposition'] = 'attachment; filename=LPTL_Transer_Report.xlsx'
                                                wb.save(response)
                                                return response

        # Case 04: if detchanges file is empty. Copy sumall/summary as final report

        elif allBilled == False and someBilled == True:
            # source_file = wb4
            # source_sheet_name = "Sheet1"
            # output_file = wb

            # source_wb = openpyxl.load_workbook(source_file)
            source_ws = ws4
            # output_wb = openpyxl.Workbook()
            output_ws = ws
            # output_ws.title = source_sheet_name

            for row in source_ws.iter_rows():
                for cell in row:
                    output_ws[cell.coordinate].value = cell.value
                    output_ws[cell.coordinate].font = cell.font.copy()
                    output_ws[cell.coordinate].border = cell.border.copy()
                    output_ws[cell.coordinate].fill = cell.fill.copy()
                    output_ws[cell.coordinate].number_format = cell.number_format
                    output_ws[cell.coordinate].protection = cell.protection.copy()
                    output_ws[cell.coordinate].alignment = cell.alignment.copy()
                    output_ws[cell.coordinate].comment = cell.comment
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=LPTL_Transer_Report.xlsx'
            wb.save(response)
            return response




        # comparing current months samplesresults file with billing detail all file to know if the already billed sample is present in current month's billing detail all file. If yes, subtract it from final report
        # prevs1=None
        # prevt1=None
        # for rowsam in range(2, ws7.max_row):
        #         if ws7['AE' + str(rowsam)].value == "Y":
        #             s1 = ws7['C' + str(rowsam)].value
        #             for rowbildall in range(2,ws2.max_row):
        #                 s2 = ws2['C' + str(rowbildall)].value
        #                 if s1==s2:
        #                     t1 = ws7['AC' + str(rowsam)].value
        #                     t2 = ws2['G' + str(rowbildall)].value
        #                     if t1 == t2:
        #                             print(s1+t1)
        #                             print(s2+t2)
        #                             print("same sample and test code")
        #                             billingAmount=ws2['K' + str(rowbildall)].value
        #                             costCenter=ws2['B' + str(rowbildall)].value
                                    
        #                             for rowfinal in range(2,output_ws.max_row):
        #                                 if prevt1!=t1 and prevs1!=s1:
        #                                     if output_ws['D'+str(rowfinal)].value == costCenter:
        #                                         afterSubtract = output_ws['E'+str(rowfinal)].value - billingAmount
        #                                         output_ws['E'+str(rowfinal)].value = afterSubtract
        #                                         output_wb.save(filepath)
        #                             prevs1=s1
        #                             prevt1=t1